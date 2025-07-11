import asyncio
import datetime
import os
import requests
import json

from aiogram import Bot, Dispatcher, F
from aiogram.enums import ParseMode
from aiogram.types import Message, ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.client.default import DefaultBotProperties

from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import gspread
from openpyxl import Workbook, load_workbook

# === НАСТРОЙКИ ===
API_TOKEN = os.getenv("BOT_TOKEN")
HR_TELEGRAM_ID = int(os.getenv("HR_TELEGRAM_ID"))
GOOGLE_SHEET_ID = os.getenv("GOOGLE_SHEET_ID")
SHEET_NAME = os.getenv("SHEET_NAME", "Лист1")
DRIVE_FOLDER_ID = os.getenv("DRIVE_FOLDER_ID")

CREDENTIALS_FILE = "service_account.json"
with open(CREDENTIALS_FILE, "w") as f:
    f.write(os.getenv("GOOGLE_CREDENTIALS_JSON"))

bot = Bot(token=API_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher(storage=MemoryStorage())

class Form(StatesGroup):
    fio = State()
    positions = State()
    resume = State()
    contacts = State()
    consent = State()
    post_consent = State()

def yes_no_keyboard():
    return ReplyKeyboardMarkup(resize_keyboard=True, keyboard=[
        [KeyboardButton(text="Да"), KeyboardButton(text="Нет")]
    ])

def consent_keyboard():
    return ReplyKeyboardMarkup(resize_keyboard=True, keyboard=[
        [KeyboardButton(text="Даю согласие")],
        [KeyboardButton(text="Удалите мои данные")]
    ])

def authorize_google():
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scopes)
    client = gspread.authorize(creds)
    drive_service = build('drive', 'v3', credentials=creds)
    return drive_service, client

def upload_resume_to_drive(drive_service, file_id, fio):
    bot_file = asyncio.run(bot.get_file(file_id))
    file_url = f"https://api.telegram.org/file/bot{API_TOKEN}/{bot_file.file_path}"
    ext = os.path.splitext(bot_file.file_path)[1] or ".pdf"
    safe_filename = f"{fio}{ext}"
    local_path = safe_filename

    response = requests.get(file_url)
    with open(local_path, "wb") as f:
        f.write(response.content)

    file_metadata = {
        'name': safe_filename,
        'parents': [DRIVE_FOLDER_ID]
    }
    media = MediaFileUpload(local_path, resumable=True)
    uploaded_file = drive_service.files().create(
        body=file_metadata,
        media_body=media,
        fields='id'
    ).execute()

    os.remove(local_path)

    return f"https://drive.google.com/file/d/{uploaded_file['id']}/view?usp=sharing"

def write_to_google_sheets(fio, positions, contacts, resume_link, timestamp):
    drive_service, client = authorize_google()
    sheet = client.open_by_key(GOOGLE_SHEET_ID).worksheet(SHEET_NAME)
    sheet.append_row([timestamp, fio, positions, contacts, resume_link])
    print("✅ Заявка записана в Google Таблицу")

def write_to_excel(fio, positions, contacts, resume_link, timestamp):
    filename = "log.xlsx"
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append(["Время", "ФИО", "Должности", "Контакты", "Резюме (ссылка)"])
        wb.save(filename)
    wb = load_workbook(filename)
    ws = wb.active
    ws.append([timestamp, fio, positions, contacts, resume_link])
    wb.save(filename)

@dp.message(Command("start"))
async def start(message: Message, state: FSMContext):
    await state.set_state(Form.fio)
    await message.answer(
        "👋 Здравствуйте!\n"
        "Я автоматический бот, предназначенный для 📥 сбора резерва соискателей в компанию «Все Пороги».\n\n"
        "🤖 Я не умею вести диалог, поэтому, к сожалению, не смогу ответить на Ваши вопросы.\n"
        "Но обязательно помогу Вашей кандидатуре не потеряться в потоке других резюме.\n\n"
        "📝 Для этого, пожалуйста, введите Ваше ФИО:"
    )

@dp.message(Form.fio)
async def handle_fio(message: Message, state: FSMContext):
    await state.update_data(fio=message.text)
    await state.set_state(Form.positions)
    await message.answer("Спасибо! Теперь укажите предыдущие должности (через запятую):")

@dp.message(Form.positions)
async def handle_positions(message: Message, state: FSMContext):
    await state.update_data(positions=message.text)
    await state.set_state(Form.resume)
    await message.answer("Хотите прикрепить файл с резюме? Пришлите документ или напишите любой текст, чтобы пропустить.")

@dp.message(Form.resume, F.document)
async def handle_resume_file(message: Message, state: FSMContext):
    await state.update_data(resume=message.document.file_id)
    await state.set_state(Form.contacts)
    await message.answer("Оставьте контактную информацию для связи (оптимально — телефон с привязанным WhatsApp):")

@dp.message(Form.resume)
async def skip_resume(message: Message, state: FSMContext):
    await state.update_data(resume=None)
    await state.set_state(Form.contacts)
    await message.answer("Оставьте контактную информацию для связи (оптимально — телефон с привязанным WhatsApp):")

@dp.message(Form.contacts)
async def handle_contacts(message: Message, state: FSMContext):
    await state.update_data(contacts=message.text)
    await state.set_state(Form.consent)
    await message.answer("Вы даёте согласие на обработку и хранение персональных данных?", reply_markup=yes_no_keyboard())

@dp.message(Form.consent, F.text.lower() == "да")
async def handle_consent_yes(message: Message, state: FSMContext):
    await finalize(message, state)

@dp.message(Form.consent, F.text.lower() == "нет")
async def handle_consent_no(message: Message, state: FSMContext):
    await state.set_state(Form.post_consent)
    await message.answer(
        "К сожалению, по закону без согласия на обработку и хранение персональных данных мы не сможем продолжить работу с предоставленной Вами информацией. Пожалуйста, выберите следующий шаг.",
        reply_markup=consent_keyboard()
    )

@dp.message(Form.post_consent, F.text.lower() == "даю согласие")
async def handle_post_consent_yes(message: Message, state: FSMContext):
    await finalize(message, state)

@dp.message(Form.post_consent, F.text.lower() == "удалите мои данные")
async def handle_post_consent_delete(message: Message, state: FSMContext):
    await message.answer(
        "Ваши данные не были сохранены. Вы можете вручную удалить переписку, если желаете. "
        "Мы в любом случае благодарны за Ваш интерес и надеемся посотрудничать в будущем!",
        reply_markup=ReplyKeyboardRemove()
    )
    await state.clear()

async def finalize(message: Message, state: FSMContext):
    data = await state.get_data()
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    resume_link = "не предоставлено"
    if data.get("resume"):
        try:
            drive_service, _ = authorize_google()
            resume_link = upload_resume_to_drive(drive_service, data["resume"], data["fio"])
        except Exception as e:
            print("❌ Ошибка загрузки файла на Google Диск:", e)

    summary = (
        f"📥 <b>Новая заявка от соискателя</b>\n\n"
        f"👤 <b>ФИО:</b> {data.get('fio')}\n"
        f"💼 <b>Должности:</b> {data.get('positions')}\n"
        f"📞 <b>Контакты:</b> {data.get('contacts')}\n"
        f"📎 <b>Резюме:</b> {resume_link}\n"
        f"🕒 <b>Время:</b> {now}\n"
        f"✅ <b>Согласие получено</b>"
    )

    try:
        write_to_excel(data["fio"], data["positions"], data["contacts"], resume_link, now)
        write_to_google_sheets(data["fio"], data["positions"], data["contacts"], resume_link, now)
    except Exception as e:
        print("❌ Ошибка при сохранении в таблицы:", e)

    try:
        await bot.send_message(chat_id=HR_TELEGRAM_ID, text=summary)
    except Exception as e:
        print("❌ Ошибка отправки сообщения HR:", e)

    await message.answer(
        "Спасибо! Ваши данные успешно добавлены в резерв. Мы свяжемся, как только появится подходящая для Вас вакансия.",
        reply_markup=ReplyKeyboardRemove()
    )
    await state.clear()

async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
